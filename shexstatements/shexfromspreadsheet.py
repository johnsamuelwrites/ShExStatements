#
# SPDX-FileCopyrightText: 2020 John Samuel <johnsamuelwrites@gmail.com>
#
# SPDX-License-Identifier: GPL-3.0-or-later
#

from os import remove
from os.path import splitext
import tempfile

from odf.opendocument import load
from odf.table import TableCell, TableRow
from openpyxl import load_workbook
from xlrd import open_workbook

from shexstatements.shexfromcsv import CSV


class Spreadsheet:
    """
      This class contains functions that can be used to generate ShEx from a data string or a Spreadsheet file
      (.xlsx,.xls,.ods)
    """

    @staticmethod
    def generate_shex_from_spreadsheet(filepath, skip_header=False, stream=None):
        """
        This method can be used to generate ShEx from data string. However, the input data string must contain one or more lines. Each line contains '|' separated values. If filepath is a string, filename  should be set to false.

        Parameters
        ----------
          filepath : str
            This parameter contains path of a Spreadsheet file
          skip_header : bool
            if the first line is a header, set this value to True. By default, the value is False.

        Returns
        -------
          shex
            shape expression

        """
        shexstatement = ""
        try:
            data = ""
            filename, file_extension = splitext(filepath)

            if(file_extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}):
                wb = None
                temp_path = None
                if stream is not None:
                    fd, temp_path = tempfile.mkstemp(suffix=file_extension)
                    try:
                        with open(fd, "wb") as sf:
                            sf.write(stream)
                    except TypeError:
                        # Fallback for environments where opening by fd is not supported
                        import os
                        os.close(fd)
                        with open(temp_path, "wb") as sf:
                            sf.write(stream)
                    filepath_to_open = temp_path
                else:
                    filepath_to_open = filepath

                wb = load_workbook(filepath_to_open)
                for ws in wb.worksheets:
                    for i in range(1, ws.max_row+1):
                        line = list()
                        for j in range(1, ws.max_column+1):
                            cell = ws.cell(row=i, column=j).value
                            if cell is not None:
                                line.append(cell)
                        line = "|".join(line)
                        data = data + line + "\n"

                if stream is not None and temp_path is not None:
                    remove(temp_path)

            elif(file_extension in {".xls"}):
                wb = None
                if stream is not None:
                    #wb = open_workbook(file_contents=stream, encoding_override="cp1252")
                    wb = open_workbook(file_contents=stream)
                else:
                    wb = open_workbook(filepath)
                for sheet in wb.sheets():
                    for i in range(0, wb.sheets()[0].nrows):
                        line = list()
                        for j in range(0, wb.sheets()[0].ncols):
                            cell = sheet.cell(i, j).value
                            if len(str(cell)) > 0:
                                line.append(cell)
                        data = data + "|".join(line) + "\n"

            elif(file_extension in {".ods"}):
                wb = None
                temp_path = None
                if stream is not None:
                    fd, temp_path = tempfile.mkstemp(suffix=file_extension)
                    try:
                        with open(fd, "wb") as sf:
                            sf.write(stream)
                    except TypeError:
                        # Fallback for environments where opening by fd is not supported
                        import os
                        os.close(fd)
                        with open(temp_path, "wb") as sf:
                            sf.write(stream)
                    filepath_to_open = temp_path
                else:
                    filepath_to_open = filepath

                wb = load(filepath_to_open)
                wb = wb.spreadsheet
                rows = wb.getElementsByType(TableRow)
                for row in rows:
                    cells = row.getElementsByType(TableCell)
                    line = list()
                    for cell in cells:
                        if len(str(cell)) > 0:
                            line.append(str(cell))
                    data = data+"|".join(line) + "\n"

                if stream is not None and temp_path is not None:
                    remove(temp_path)

            shexstatement = CSV.generate_shex_from_data_string(data)
        except Exception as e:
            print("Unable to read file. Error: " + str(e))
        return shexstatement
